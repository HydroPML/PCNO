import datetime
import os
import random
import numpy as np
import torch.nn.functional as F

os.environ["CUDA_VISIBLE_DEVICES"]="0" # Select the available GPU
from models.oned_unet import Unet
from models.PCNO1D import PCNO1d

from PIL import Image
import imageio
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.colors as mcolors
import colorsys

from functools import partial
from KS.utils_ks import pde_data, LpLoss, eq_check_rt, eq_check_rf, kerras_boundaries # utils_ks is for 1D KSE with varying viscosity
# from KS.utils import pde_data, LpLoss, eq_check_rt, eq_check_rf, kerras_boundaries # utils is for 1D KSE with fixed viscosity
from loss import CustomMSELoss, PearsonCorrelationScore, ScaledLpLoss
from ema import ExponentialMovingAverage
from diffusers.schedulers import DDPMScheduler
from matplotlib.colors import ListedColormap
import imageio
# from data.datamodule import PDEDataModule

import math
import scipy
import numpy as np
from timeit import default_timer
import argparse
from torch.utils.tensorboard import SummaryWriter
import torch
import torch.nn as nn
import h5py
import xarray as xr
from tqdm import tqdm
from torch.cuda.amp import autocast, GradScaler

from models.diffusion import ElucidatedDiffusion
from openpyxl import load_workbook

torch.set_num_threads(1)



################################################################
# configs
################################################################

parser = argparse.ArgumentParser()

parser.add_argument("--results_path", type=str, default="/Path_to/results/KDE/DiffPCNO/", help="path to store results")
parser.add_argument("--suffix", type=str, default="seed1", help="suffix to add to the results path")
parser.add_argument("--txt_suffix", type=str, default="1D_KDE_DiffPCNO_sampling_seed1", help="suffix to add to the results txt")
parser.add_argument("--super", type=str, default='False', help="enable superres testing")
parser.add_argument("--verbose",type=str, default='True')

parser.add_argument("--T", type=int, default=20, help="number of timesteps to predict")
parser.add_argument("--ntrain", type=int, default=1000, help="training sample size")
parser.add_argument("--nvalid", type=int, default=100, help="valid sample size")
parser.add_argument("--ntest", type=int, default=100, help="test sample size")
parser.add_argument("--nsuper", type=int, default=None)
parser.add_argument("--seed", type=int, default=1)

parser.add_argument("--model_type", type=str, default='Unet')
parser.add_argument("--depth", type=int, default=4)
parser.add_argument("--modes", type=int, default=12)
parser.add_argument("--width", type=int, default=20)
parser.add_argument("--Gwidth", type=int, default=10, help="hidden dimension of equivariant layers if model_type=hybrid")
parser.add_argument("--n_equiv", type=int, default=3, help="number of equivariant layers if model_type=hybrid")
parser.add_argument("--reflection", action="store_true", help="symmetry group p4->p4m for data augmentation")
parser.add_argument("--grid", type=str, default=None, help="[symmetric, cartesian, None]")

parser.add_argument("--epochs", type=int, default=100)
parser.add_argument("--early_stopping", type=int, default=100, help="stop if validation error does not improve for successive epochs")
parser.add_argument("--batch_size", type=int, default=128)
parser.add_argument("--learning_rate", type=float, default=0.0001)
parser.add_argument("--step", action="store_true", help="use step scheduler")
parser.add_argument("--gamma", type=float, default=None, help="gamma for step scheduler")
parser.add_argument("--step_size", type=int, default=None, help="step size for step scheduler")
parser.add_argument("--lmbda", type=float, default=0.0001, help="weight decay for adam")
parser.add_argument("--strategy", type=str, default="markov", help="markov, recurrent or oneshot")
parser.add_argument("--time_pad", action="store_true", help="pad the time dimension for strategy=oneshot")
parser.add_argument("--noise_std", type=float, default=0.00, help="amount of noise to inject for strategy=markov")

# Model parameters
parser.add_argument("--max_num_steps", type=int, default=1000, help="Maximum number of steps")
parser.add_argument("--criterion", type=str, default="mse", help="Loss criterion")
parser.add_argument("--param_conditioning", type=str, default="scalar_3", help="Parameter conditioning for 1D KSE with varying viscosity")
# parser.add_argument("--param_conditioning", type=str, default="scalar_2", help="Parameter conditioning for 1D KSE with fixed viscosity")
parser.add_argument("--padding_mode", type=str, default="circular", help="Padding mode")
parser.add_argument("--predict_difference", type=bool, default=True, help="Predict difference flag")
parser.add_argument("--difference_weight", type=float, default=0.3, help="Difference weight")
parser.add_argument("--min_noise_std", type=float, default=4e-7, help="Minimum noise standard deviation")
parser.add_argument("--ema_decay", type=float, default=0.995, help="Minimum noise standard deviation")
parser.add_argument("--num_refinement_steps", type=int, default=3, help="Number of refinement steps")
parser.add_argument("--time_history", type=int, default=1, help="Time history steps")
parser.add_argument("--time_future", type=int, default=1, help="Time future steps")
parser.add_argument("--time_gap", type=int, default=0, help="Time gap")

# Data PDE parameters
parser.add_argument("--n_scalar_components", type=int, default=1, help="Number of scalar components in PDE")
parser.add_argument("--n_vector_components", type=int, default=0, help="Number of vector components in PDE")
parser.add_argument("--trajlen", type=int, default=140, help="Trajectory length")
parser.add_argument("--n_spatial_dim", type=int, default=1, help="Number of spatial dimensions")
parser.add_argument("--activation", type=str, default="gelu", help="activation")

parser.add_argument("--train_limit_trajectories", type=int, default=-1, help="Train limit trajectories")
parser.add_argument("--valid_limit_trajectories", type=int, default=-1, help="Validation limit trajectories")
parser.add_argument("--test_limit_trajectories", type=int, default=-1, help="Test limit trajectories")

args = parser.parse_args()

assert args.model_type in ["Unet"], f"Invalid model type {args.model_type}"
assert args.strategy in ["teacher_forcing", "markov", "recurrent", "oneshot"], "Invalid training strategy"

torch.manual_seed(args.seed)
np.random.seed(args.seed)
torch.cuda.manual_seed(args.seed)
random.seed(args.seed)
def set_random_seed(seed):
    torch.manual_seed(seed)
    np.random.seed(seed)
    torch.cuda.manual_seed(seed)
    random.seed(seed)

data_aug = "aug" in args.model_type


# FNO data specs
S = 256 # spatial res
S_super = 4 * S # super spatial res
T_in = 1 # number of input times
T_train = 139
T_valid = 399
T_test = 399
d = 2 # spatial res
num_channels = 1



# adjust data specs based on model type and data path
threeD = args.model_type in ["FNO3d", "FNO3d_aug",
                             "GCNN3d_p4", "GCNN3d_p4m",
                             "GFNO3d_p4", "GFNO3d_p4m",
                             "radialNO3d_p4", "radialNO3d_p4m",
                             "PCNO3d"]

grid_type = "symmetric"
swe = False
if args.grid:
    grid_type = args.grid
    assert grid_type in ['symmetric', 'cartesian', 'None']

spatial_dims = range(1, d + 1)

if args.strategy == "oneshot":
    assert threeD, "oneshot strategy only for 3d models"

if threeD:
    assert args.strategy == "oneshot", "threeD models use oneshot strategy"
    # assert args.modes <= 8, "modes for 3d models should be leq 8"

ntrain = args.ntrain # 1000
nvalid = args.nvalid
ntest = args.ntest # 100

time_modes = None
time = args.strategy == "oneshot" # perform convolutions in space-time
if time and not args.time_pad:
    time_modes = 5 if swe else 6 # 6 is based on T=10
elif time and swe:
    time_modes = 8

modes = args.modes
width = args.width
n_layer = args.depth
batch_size = args.batch_size

epochs = args.epochs # 500
learning_rate = args.learning_rate
scheduler_step = args.step_size
scheduler_gamma = args.gamma # for step scheduler

initial_step = 1 if args.strategy == "markov" else T_in

root = '/Path_to_pre_trained/Models/1D_KSE/PCNO-Refiner'
path_model = os.path.join(root, 'model.pt')
writer = SummaryWriter(root)

################################################################
# Model init
################################################################

if args.model_type in ["Unet"]:
    model = Unet(n_input_scalar_components=args.n_scalar_components, n_input_vector_components=args.n_vector_components,
         n_output_scalar_components=args.n_scalar_components,
         n_output_vector_components=args.n_vector_components, time_history=args.time_history + args.time_future,
         time_future=args.time_future, hidden_channels=64, activation=args.activation,
         norm=True, param_conditioning=args.param_conditioning, n_dims=args.n_spatial_dim).cuda()
else:
    raise NotImplementedError("Model not recognized")

ema_model = Unet(n_input_scalar_components=args.n_scalar_components, n_input_vector_components=args.n_vector_components,
         n_output_scalar_components=args.n_scalar_components,
         n_output_vector_components=args.n_vector_components, time_history=args.time_history + args.time_future,
         time_future=args.time_future, hidden_channels=64, activation=args.activation,
         norm=True, param_conditioning=args.param_conditioning, n_dims=args.n_spatial_dim).cuda()
ema_model.load_state_dict(model.state_dict())

model_con = PCNO1d(num_channels=num_channels, hidden_channels=64, param_conditioning=args.param_conditioning, modes=modes, width=width, initial_step=initial_step).cuda()
path_model_con = 'Path_to/Models/1D_KSE/PCNO/Fixed_viscosity/model.pt' # pre_trained PCNO of 1D KSE with fixed viscosity
# path_model_con = 'Path_to/Models/1D_KSE/PCNO/Varying_viscosity/model.pt' # pre_trained PCNO of 1D KSE with varying viscosity
model_con.load_state_dict(torch.load(path_model_con))
model_con.eval()
ema = ExponentialMovingAverage(model, decay=args.ema_decay)

# model.load_state_dict(torch.load(path_model))
# checkpoint = torch.load(path_model)
# # model.load_state_dict(checkpoint['model'])
# ema.shadow = checkpoint["ema"]

################################################################
# load data
################################################################
def dataset(path, mode):
    with h5py.File(path, "r") as f:
        data_h5 = f[mode]
        data_key = [k for k in data_h5.keys() if k.startswith("pde_")][0]
        data = {
            "u": torch.tensor(data_h5[data_key][:].astype(np.float32)),
            "dt": torch.tensor(data_h5["dt"][:].astype(np.float32)),
            "dx": torch.tensor(data_h5["dx"][:].astype(np.float32)),
        }
        # data_u = data["u"].permute(0, 2, 1)

        # print('data["u"]', data["u"].shape)
        # print('data["dt"]', data["dt"].shape)
        # print('data["dx"]', data["dx"].shape)
        if "v" in data_h5:
            data["v"] = torch.tensor(data_h5["v"][:].astype(np.float32))

        data["orig_dt"] = data["dt"].clone()
        if data["u"].ndim == 3:
            data["u"] = data["u"].unsqueeze(dim=-2)  # Add channel dimension

        # Normalizing the data to [0,1]
        shift_u = data["u"].min()
        scale_u = data["u"].max() - data["u"].min()
        data["u"] = (data["u"] - shift_u) / scale_u
        # data["u"] = 2 * data["u"] - 1
        data["u"] = (data["u"] - 0.5) / 0.5

        # The KS equation is parameterized by [1] the time step between observations
        # (measured in seconds, usually around 0.2), [2] the spatial step between
        # data points in the spatial domain (measured in meters, usually around 0.2),
        # and finally [3] the viscosity parameter (measured in m^2/s, usually between 0.5 - 1.5).
        # We scale these parameters to be in the range [0, 10] to be visible changes in fourier embeds.
        # This accelerates learning and makes it easier for the models to learn the conditional dynamics.
        # # Scaling time step.
        if data["dt"].min() > 0.15 and data["dt"].max() < 0.25:
            data["dt"] = (data["dt"] - 0.15) * 100.0
        else:
            print(
                f"WARNING: dt is not in the expected range (min {data['dt'].min()}, max {data['dt'].max()}, mean {data['dt'].mean()}) - scaling may be incorrect."
            )
        # Scaling spatial step.
        if data["dx"].min() > 0.2 and data["dx"].max() < 0.3:
            data["dx"] = (data["dx"] - 0.2) * 100.0
        else:
            print(
                f"WARNING: dx is not in the expected range (min {data['dx'].min()}, max {data['dx'].max()}, mean {data['dx'].mean()}) - scaling may be incorrect."
            )
        # Scaling viscosity.
        if "v" in data:
            if data["v"].min() >= 0.5 and data["v"].max() <= 1.5:
                data["v"] = (data["v"] - 0.5) * 100.0
            else:
                print(
                    f"WARNING: v is not in the expected range (min {data['v'].min()}, max {data['v'].max()}, mean {data['v'].mean()}) - scaling may be incorrect."
                )
        # print('data["u"]', data["u"].shape)
        # print('data["dt"]', data["dt"].shape)
        # print('data["dx"]', data["dx"].shape)
    return data

# 1D KSE with varying viscosity
Path_train = '/Path_to/Dataset/KS/KS_train_conditional_viscosity.h5'
Path_valid = '/Path_to/Dataset/KS/KS_valid_conditional_viscosity.h5'
Path_test = '/Path_to/Dataset/KS/KS_test_conditional_viscosity.h5'
# # 1D KSE with fixed viscosity
# Path_train = '/Path_to/Dataset/KS/KS_train_fixed_viscosity.h5'
# Path_valid = '/Path_to/Dataset/KS/KS_valid_fixed_viscosity.h5'
# Path_test = '/Path_to/Dataset/KS/KS_test_fixed_viscosity.h5'

train = dataset(path=Path_train, mode='train')
test = dataset(path=Path_test, mode='test')
valid = dataset(path=Path_valid, mode='valid')

# if args.verbose:
# print(f"{args.model_type}: Train/valid/test data shape: ")
# print(train.shape)
# print(valid.shape)
# print(test.shape)

train_data = pde_data(train, train=True, strategy=args.strategy, T_in=T_in, T_out=T_train, std=args.noise_std)
ntrain = len(train_data)
valid_data = pde_data(valid, train=False, strategy=args.strategy, T_in=T_in, T_out=T_valid)
nvalid = len(valid_data)
test_data = pde_data(test, train=False, strategy=args.strategy, T_in=T_in, T_out=T_test)
ntest = len(test_data)
# print('ntrain',ntrain)
print('nvalid',nvalid)
print('ntest',ntest)

train_loader = torch.utils.data.DataLoader(train_data, batch_size=batch_size, shuffle=True)
valid_loader = torch.utils.data.DataLoader(valid_data, batch_size=batch_size, shuffle=False)
test_loader = torch.utils.data.DataLoader(test_data, batch_size=batch_size, shuffle=False)

################################################################
# training and evaluation
################################################################


batch_std = []
for xx, yy, cond in train_loader:
    batch_std.append(yy.std())
batch_std_tensor = torch.stack(batch_std)
sigma_data = batch_std_tensor.mean()

# model = ElucidatedDiffusion(net, channels = num_channels, image_size=S, sigma_data=sigma_data).cuda()
# scaler = GradScaler()

complex_ct = sum(par.numel() * (1 + par.is_complex()) for par in model.parameters())
real_ct = sum(par.numel() for par in model.parameters())
if args.verbose:
    print(f"{args.model_type}; # Params: complex count {complex_ct}, real count: {real_ct}")
writer.add_scalar("Parameters/Complex", complex_ct)
writer.add_scalar("Parameters/Real", real_ct)

# optimizer = torch.optim.Adam(model.parameters(), lr=learning_rate, weight_decay=args.lmbda)
optimizer = torch.optim.RAdam(model.parameters(), lr=learning_rate, weight_decay=0.0)
if args.step:
    assert args.step_size is not None, "step_size is None"
    assert scheduler_gamma is not None, "gamma is None"
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=args.step_size, gamma=scheduler_gamma)
else:
    num_training_steps = epochs * len(train_loader)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=num_training_steps)

lploss = LpLoss(size_average=False)

best_valid = float("inf")



if (args.n_spatial_dim) == 3:
    _mode = "3D"
    nn.Conv3d = partial(nn.Conv3d, padding_mode=args.padding_mode)
elif (args.n_spatial_dim) == 2:
    _mode = "2D"
    nn.Conv2d = partial(nn.Conv2d, padding_mode=args.padding_mode)
elif (args.n_spatial_dim) == 1:
    _mode = "1D"
    nn.Conv1d = partial(nn.Conv1d, padding_mode=args.padding_mode)
else:
    raise NotImplementedError(f"{pde}")
train_criterion = CustomMSELoss()


# implement the denoising manually.
betas = [args.min_noise_std ** (k / args.num_refinement_steps) for k in reversed(range(args.num_refinement_steps + 1))]
scheduler_DDPM = DDPMScheduler(
    num_train_timesteps=args.num_refinement_steps + 1,
    trained_betas=betas,
    prediction_type="v_prediction",
    clip_sample=False,
)
# Multiplies k before passing to frequency embedding.
time_multiplier = 1000 / args.num_refinement_steps

val_criterions = {"mse": CustomMSELoss(), "scaledl2": ScaledLpLoss()}
rollout_criterions = {"mse": torch.nn.MSELoss(reduction="none"), "corr": PearsonCorrelationScore()}
time_resolution = args.trajlen
# Max number of previous points solver can eat
reduced_time_resolution = time_resolution - args.time_history
        # Number of future points to predict
max_start_time = (reduced_time_resolution - args.time_future * args.max_num_steps - args.time_gap)
max_start_time = max(0, max_start_time)

def predict_next_solution(model, x, cond):
    y = model.sample(
        (x[:,:1,:,:] - 0.5) / 0.5 + torch.randn_like(x[:,1:,:,:]).cuda() * 80.0,
        list(reversed([0.661, 0.9, 5.84, 24.4, 80.0])),
        x,
        z=cond,
    )
    # if args.predict_difference:
    #     y = y * args.difference_weight + x[:,1:,:,:]

    y = (y * 0.5 + 0.5).clamp(0, 1)
    return y

def get_eval_pred_m(model, model_con, x, cond, strategy, T, times, num_samples=2):
    all_preds = []

    for i in range(num_samples):
        # seed = random.randint(0, 2**32 - 1)
        # set_random_seed(seed)

        pred = None
        x0 = x
        for t in range(T):
            t1 = default_timer()
            with torch.no_grad():
                im_c = model_con(x0, z=cond)
            im_cc = torch.cat((im_c, x0), dim=1)
            im = predict_next_solution(model, im_cc, cond)
            times.append(default_timer() - t1)
            if t == 0:
                pred = im
            else:
                pred = torch.cat((pred, im), 1)
            if strategy == "markov":
                x0 = im_c
            else:
                x0 = torch.cat((x0[..., 1:, :], im), dim=1)

        all_preds.append(pred.unsqueeze(0))
    all_preds = torch.cat(all_preds, dim=0)

    mean_pred = all_preds.mean(dim=0)
    std_pred = all_preds.std(dim=0)

    return mean_pred, std_pred




def apply_ema():
    ema.apply_shadow()
def remove_ema():
    ema.restore()

model.eval()

start = default_timer()
if args.verbose:
    print("Training...")
step_ct = 0
train_times = []
eval_times = []
# ema.register()
for param in model_con.parameters():
    param.requires_grad = False



# model.eval()

def generate_movie_2D(key, test_x, test_y, preds_y, preds_std, plot_title='', field=0, val_cbar_index=-1, err_cbar_index=-1,
                      val_clim=None, err_clim=None, std_clim=None, font_size=None, movie_dir='', movie_name='movie.gif',
                      frame_basename='movie', frame_ext='jpg', remove_frames=True):
    frame_files = []

    if movie_dir:
        os.makedirs(movie_dir, exist_ok=True)

    if font_size is not None:
        plt.rcParams.update({'font.size': font_size})

    if len(preds_y.shape) == 4:
        Nsamples, Nx, Ny, Nt = preds_y.shape
        preds_y = preds_y.reshape(Nsamples, Nx, Ny, Nt, 1)
        test_y = test_y.reshape(Nsamples, Nx, Ny, Nt, 1)
        preds_std = preds_std.reshape(Nsamples, Nx, Ny, Nt, 1)
    Nsamples, Nx, Ny, Nt, Nfields = preds_y.shape
    print('preds_y', preds_y.shape)

    pred = preds_y[key, ..., field]
    true = test_y[key, ..., field]
    std = preds_std[key, ..., field]
    error = torch.abs(pred - true)

    a = test_x[key]
    x = torch.linspace(0, 1, Nx + 1)[:-1]
    y = torch.linspace(0, 1, Ny + 1)[:-1]
    X, Y = torch.meshgrid(x, y)
    # print('X', X.shape)
    # print('Y', Y.shape)

    # t = a[0, 0, :, 2]
    dpi = 100
    #
    fig_width_inch = Nx / dpi
    fig_height_inch = Ny / dpi

    fig, axs = plt.subplots(1, 4, figsize=(30, 4))
    ax1 = axs[0]
    ax2 = axs[1]
    ax3 = axs[2]
    ax4 = axs[3]
    colors = plt.cm.viridis(np.linspace(0, 1, 256))
    colors[0] = [1, 1, 1, 1]
    cmap = ListedColormap(colors)

    pcm1 = ax1.pcolormesh(X, Y, true[..., val_cbar_index], cmap='RdBu_r', label='true', shading='gouraud')
    pcm2 = ax2.pcolormesh(X, Y, pred[..., val_cbar_index], cmap='RdBu_r', label='pred', shading='gouraud')
    pcm3 = ax3.pcolormesh(X, Y, error[..., err_cbar_index], cmap='RdBu_r', label='error', shading='gouraud')
    pcm4 = ax4.pcolormesh(X, Y, std[..., err_cbar_index], cmap='RdBu_r', label='uncertainty', shading='gouraud')

    if val_clim is None:
        val_clim = pcm1.get_clim()
    if err_clim is None:
        err_clim = pcm3.get_clim()
    if std_clim is None:
        std_clim = pcm4.get_clim()

    pcm1.set_clim(val_clim)
    plt.colorbar(pcm1, ax=ax1)
    ax1.set_aspect('auto')
    # ax1.axis('square')

    pcm2.set_clim(val_clim)
    plt.colorbar(pcm2, ax=ax2)
    ax2.set_aspect('auto')
    # ax2.axis('square')

    pcm3.set_clim(err_clim)
    plt.colorbar(pcm3, ax=ax3)
    ax3.set_aspect('auto')
    # ax3.axis('square')

    pcm4.set_clim(std_clim)
    plt.colorbar(pcm4, ax=ax4)
    ax4.set_aspect('auto')

    # cmap = plt.cm.RdBu
    # def desaturate_cmap(cmap, factor=0.5):
    #     new_cmap = cmap(np.arange(cmap.N))
    #     for i in range(new_cmap.shape[0]):
    #         new_cmap[i, :3] = new_cmap[i, :3] * factor + (1 - factor) * 0.5
    #     return mcolors.ListedColormap(new_cmap)
    # desaturated_cmap = desaturate_cmap(cmap, factor=0.2)

    plt.tight_layout()
    # data_list = [
    #     (true, "Exact"),
    #     (pred, "Predictions"),
    #     (error, "Errors"),
    #     (std, "Uncertainty"),
    # ]
    #
    # fig_width = Nx
    # fig_height = Ny
    # #
    # for i in range(Nt):
    #     for data, name in data_list:
    #         plt.figure()
    #         plt.pcolormesh(X, Y, data[..., i], cmap="RdBu_r", shading="gouraud")
    #         plt.axis("off")
    #         plt.gca().set_aspect("auto")
    #         plt.xlim(X.min(), X.max())
    #         plt.ylim(Y.min(), Y.max())
    #         ax1.set_axis_off()
    #
    #         plt.show()

    for i in range(Nt):
        # Exact
        ax1.clear()
        pcm1 = ax1.pcolormesh(X, Y, true[..., i], cmap='RdBu_r', label='true', shading='gouraud')
        pcm1.set_clim(val_clim)
        ax1.set_aspect('auto')
        ax1.set_xlim(X.min(), X.max())
        ax1.set_ylim(Y.min(), Y.max())
        ax1.set_axis_off()
        # ax1.set_title(f'Ground truth {plot_title}')
        # ax1.axis('square')

        # Predictions
        ax2.clear()
        pcm2 = ax2.pcolormesh(X, Y, pred[..., i], cmap='RdBu_r', label='pred', shading='gouraud')
        pcm2.set_clim(val_clim)
        # ax2.set_title(f'Case 1 (DNO-3) {plot_title}')
        ax2.set_aspect('auto')
        ax2.set_xlim(X.min(), X.max())
        ax2.set_ylim(Y.min(), Y.max())
        ax2.set_axis_off()
        # ax2.axis('square')

        # Error
        ax3.clear()
        pcm3 = ax3.pcolormesh(X, Y, error[..., i], cmap='RdBu_r', label='error', shading='gouraud')
        pcm3.set_clim(err_clim)
        # ax3.set_title(f'Errors {plot_title}')
        ax3.set_aspect('auto')
        ax3.set_xlim(X.min(), X.max())
        ax3.set_ylim(Y.min(), Y.max())
        ax3.set_axis_off()
        # ax3.axis('square')

        # Uncertainty
        ax4.clear()
        pcm4 = ax4.pcolormesh(X, Y, std[..., i], cmap='RdBu_r', label='uncertainty', shading='gouraud')
        pcm4.set_clim(std_clim)
        ax4.set_aspect('auto')
        ax4.set_xlim(X.min(), X.max())
        ax4.set_ylim(Y.min(), Y.max())
        ax4.set_axis_off()

        #         plt.tight_layout()
        # fig.canvas.draw()
        #
        if movie_dir:
            frame_path = os.path.join(movie_dir, f'{frame_basename}-{i:03}.{frame_ext}')
            frame_files.append(frame_path)
            plt.savefig(frame_path, dpi=900, bbox_inches='tight')
        # plt.draw()
        # plt.pause(0.1)
    #
    # if movie_dir:
    #     movie_path = os.path.join(movie_dir, movie_name)
    #     with imageio.get_writer(movie_path, mode='I') as writer:
    #         for frame in frame_files:
    #             image = imageio.imread(frame)
    #             writer.append_data(image)
    # plt.show()

    # if movie_dir and remove_frames:
    #     for frame in frame_files:
    #         try:
    #             os.remove(frame)
    #         except:
    #             pass


model.load_state_dict(torch.load(path_model))
test_l2 = test_vort_l2 = test_pres_l2 = 0
rotations_l2 = 0
reflections_l2 = 0
test_rt_l2 = 0
test_rf_l2 = 0
test_loss_by_channel = None
key = 0
i = 0
# apply_ema()
with torch.no_grad():
    model.eval()
    for xx, yy, cond in test_loader:
        x = xx.cuda()
        y = yy.cuda()
        x = x * 0.5 + 0.5
        y = y * 0.5 + 0.5
        input_data = torch.zeros_like(y).cuda()
        input_data = input_data.permute(0, 3, 1, 2)
        print('input_data',input_data.shape)
        cond = cond.cuda()
        im = model_con(x, z=cond)
        # im_c = torch.cat((im, x), dim=1)

        pred, pred_std = get_eval_pred_m(model=model, model_con=model_con, x=x, cond=cond, strategy=args.strategy, T=y.shape[1], times=[], num_samples=50)
        print('pred', pred.shape)
        print('pred_std', pred_std.shape)
        test_l2 += lploss(pred.reshape(len(pred), -1, num_channels), y.reshape(len(y), -1, num_channels)).item()
        #
        y_hm = y.permute(0, 3, 1, 2)
        pred_hm = pred.permute(0, 3, 1, 2)
        pred_std_hm = pred_std.permute(0, 3, 1, 2)

        gt_f = y_hm[..., 0]
        pred_f = pred_hm[..., 0]
        pred_std_f = pred_std_hm[..., 0]
        if i == 0:
            g_f = gt_f
        else:
            g_f = torch.cat((g_f, gt_f), 0)
        print('g_f', g_f.shape)

        if i == 0:
            p_f = pred_f
        else:
            p_f = torch.cat((p_f, pred_f), 0)

        if i == 0:
            p_std_f = pred_std_f
        else:
            p_std_f = torch.cat((p_std_f, pred_std_f), 0)
        #
        # i = i + 1

        #Visulazation
        y_hmm = y.permute(0, 3, 1, 2)
        gt_m = torch.rot90(y_hmm, k=-1, dims=[1, 2])
        print('gt', gt_m.shape)

        pred_hmm = pred.permute(0, 3, 1, 2)
        outm = torch.rot90(pred_hmm, k=-1, dims=[1, 2])

        pred_std_hmm = pred_std.permute(0, 3, 1, 2)
        outm_std = torch.rot90(pred_std_hmm, k=-1, dims=[1, 2])

        print('pred', outm.shape)
        print('outm_std', outm_std.shape)
        # MOIVE
        movie_dir = '/Path_to/Results/KS/PCNO-Refiner/%s/' % (str(i))
        os.makedirs(movie_dir, exist_ok=True)
        # H
        movie_name = 'H.gif'
        frame_basename = 'H_frame'
        frame_ext = 'jpg'
        plot_title = ""
        field = 0
        val_cbar_index = -1
        err_cbar_index = -1
        font_size = 12
        remove_frames = True
        generate_movie_2D(key, input_data.cpu(), gt_m.cpu(), outm.cpu(), outm_std.cpu(),
                          plot_title=plot_title,
                          field=field,
                          val_cbar_index=val_cbar_index,
                          err_cbar_index=err_cbar_index,
                          movie_dir=movie_dir,
                          movie_name=movie_name,
                          frame_basename=frame_basename,
                          frame_ext=frame_ext,
                          remove_frames=remove_frames,
                          font_size=font_size)
        i = i+1
# remove_ema()
    # writer.add_scalar("Test/Loss", test_l2 / ntest, best_epoch)

# test_time_l2 = test_space_l2 = ntest_super = test_int_space_l2 = test_int_time_l2 = None
# error_f = (p_f - g_f) ** 2
# print('g_f', g_f.shape)
# print('p_f', p_f.shape)
# print('p_std_f', p_std_f.shape)
# torch.save(g_f, '/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/g_f.pt')
# torch.save(p_f, '/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/p_f.pt')
# torch.save(p_std_f, '/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/p_std_f.pt')
# gt_f = g_f.cpu().numpy()
# pred_f = p_f.cpu().numpy()
# pred_std_f = p_std_f.cpu().numpy()
# e_f = error_f.cpu().numpy()
# g_avg = np.mean(gt_f, axis=1)
# p_avg = np.mean(pred_f, axis=1)
# p_std_avg = np.mean(pred_std_f, axis=1)
# e_avg = np.mean(e_f, axis=1)
# file1 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/gt.xlsx"
# file2 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/pre.xlsx"
# file3 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/gt_ave.xlsx"
# file4 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/pre_ave.xlsx"
# file5 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/e_ave.xlsx"
# file6 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/pre_std.xlsx"
# file7 = "/home/yuezeng/UrbanFloodCast/NMI_PCNN/Results/KS2/Refiner/pre_std_ave.xlsx"
# #
# # # #
# g_mean = np.mean(g_avg, axis=0).reshape(-1, 1)
# p_mean = np.mean(p_avg, axis=0).reshape(-1, 1)
# p_std_mean = np.mean(p_std_avg, axis=0).reshape(-1, 1)
# e_mean = np.mean(e_avg, axis=0).reshape(-1, 1)
# #
# # # #
# df_g = pd.DataFrame(g_avg.T)
# df_g.to_excel(file1, index=False, header=False)
# df_p = pd.DataFrame(p_avg.T)
# df_p.to_excel(file2, index=False, header=False)
# df_p_std = pd.DataFrame(p_std_avg.T)
# df_p_std.to_excel(file6, index=False, header=False)
# # #
# # #
# df_gs = pd.DataFrame(g_mean, columns=['DiffPCNO-Refiner'])
# df_gs.to_excel(file3, index=False)
# df_ps = pd.DataFrame(p_mean, columns=['DiffPCNO-Refiner'])
# df_ps.to_excel(file4, index=False)
# df_es = pd.DataFrame(e_mean, columns=['DiffPCNO-Refiner'])
# df_es.to_excel(file5, index=False)
# df_pss = pd.DataFrame(p_std_mean, columns=['DiffPCNO-Refiner'])
# df_pss.to_excel(file7, index=False)
#
# # #
# # #
# wb = load_workbook(file3)
# ws = wb.active
# max_col = ws.max_column
# ws.cell(row=1, column=max_col+1, value="ICM_UNet_0.001")
# for i in range(159):
#     ws.cell(row=i+2, column=max_col+1, value=g_mean[i, 0])
# wb.save(file3)
#
# wb_p = load_workbook(file4)
# ws_p = wb_p.active
# max_col_p = ws_p.max_column
# ws_p.cell(row=1, column=max_col_p+1, value="ICM_UNet_0.001")
# for i in range(159):
#     ws_p.cell(row=i+2, column=max_col_p+1, value=p_mean[i, 0])
# wb_p.save(file4)
#
# wb_e = load_workbook(file5)
# ws_e = wb_e.active
# max_col_e = ws_e.max_column
# ws_e.cell(row=1, column=max_col_e+1, value="ICM_UNet_0.001")
# for i in range(159):
#     ws_e.cell(row=i+2, column=max_col_e+1, value=e_mean[i, 0])
# wb_e.save(file5)


print(f"{args.model_type} done training; \nTest: {test_l2 / ntest}, Rotations: {rotations_l2}, Reflections: {reflections_l2}, Super Space Test: {test_space_l2}, Super Time Test: {test_time_l2}")
writer.flush()
writer.close()